import os
from copy import copy
from src.excel.utils import get_nrows, get_names_columns

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side


def set_columns_names(ws):
    for n, col in zip(*get_names_columns()):
        ws[col][1].value = n

    ws['g'][1].fill = PatternFill(fill_type='solid', start_color="00FF00", end_color='000000')
    ws['g'][1].border = Border(left=Side(border_style='thin',
                                          color='C0C0C0'),
                                right=Side(border_style='thin',
                                           color='C0C0C0'),
                                top=Side(border_style='thin',
                                         color='C0C0C0'),
                                bottom=Side(border_style='thin',
                                            color='C0C0C0'))

    ws['k'][1].fill = PatternFill(fill_type='solid', start_color="FF0000", end_color='000000')
    ws['k'][1].border = Border(left=Side(border_style='thin',
                                         color='C0C0C0'),
                               right=Side(border_style='thin',
                                          color='C0C0C0'),
                               top=Side(border_style='thin',
                                        color='C0C0C0'),
                               bottom=Side(border_style='thin',
                                           color='C0C0C0'))


def prepare_new_excel(input_file):
    wb = load_workbook(filename=input_file)
    ws1 = wb.active

    start_col = "g"
    end_col = "z"

    # step 1
    # set summation folrmula
    # column D - column with meal's price
    nrows = get_nrows(ws1)

    for i in range(3, nrows):
        if ws1['D'][i].value is not None:
            ws1['E'][i].value = f"=SUM({start_col}{i+1}:{end_col}{i+1})"

    # step 2
    # set person's total price and colors to columns

    i = 0
    for name, col in zip(*get_names_columns()):
        ws1[col][nrows].value = f"=SUMPRODUCT({col}4:{col}{nrows}, $d4:$d{nrows})"

        i += 1
        if i % 2 == 0:
            for j in range(0, nrows+1):
                ws1[col][j].fill = PatternFill(fill_type='solid', start_color="DCDCDC", end_color='000000')
                ws1[col][j].border = Border(left=Side(border_style='thin',
                                                      color='C0C0C0'),
                                            right=Side(border_style='thin',
                                                       color='C0C0C0'),
                                            top=Side(border_style='thin',
                                                     color='C0C0C0'),
                                            bottom=Side(border_style='thin',
                                                        color='C0C0C0'))

    # step 3
    # set colors to headers
    set_columns_names(ws1)

    # step 4
    # freeze rows and columns
    ws1.freeze_panes = ws1['g'][2]

    output_file = os.path.splitext(input_file)[0] + "_.xlsx"

    wb.save(output_file)


def split_excel(input_file, output_folder):
    wb = load_workbook(filename=input_file)
    res_names = []

    for name, worksheet in zip(wb.sheetnames, wb.worksheets):
        ws1 = worksheet
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = name

        # calculate total number of rows and
        # columns in source excel file
        mr = ws1.max_row
        mc = ws1.max_column

        # copying the cell values from source
        # excel file to destination excel file
        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = ws1.cell(row=i, column=j)

                # writing the read value to destination excel file
                ws2.cell(row=i, column=j).value = c.value

                # copying cell format, including fonts, colors and etc
                if c.has_style:
                    ws2.cell(row=i, column=j).font = copy(c.font)
                    ws2.cell(row=i, column=j).border = copy(c.border)
                    ws2.cell(row=i, column=j).fill = copy(c.fill)
                    ws2.cell(row=i, column=j).number_format = copy(c.number_format)
                    ws2.cell(row=i, column=j).protection = copy(c.protection)
                    ws2.cell(row=i, column=j).alignment = copy(c.alignment)

            # after copying content, set same columns sizes
            for col in ['A', 'B', 'C', 'D']:
                ws2.column_dimensions[col].width = ws1.column_dimensions[col].width

        outname = os.path.join(output_folder, name.split()[0] + ".xlsx")
        wb2.save(outname)

        res_names.append(outname)

    return res_names


def parse_excel(input_file, output_folder):
    # todo add xls to xlsx parser!!
    res_files = split_excel(input_file, output_folder)
    # res_files = ['../../test_data/10.02.20.xlsx']

    for file in res_files:
        prepare_new_excel(file)


def test():
    input_file = '../../../2020/07_july 2020/06-10.07.2020.xlsx'
    output_folder = '../../../2020/07_july 2020/'

    parse_excel(input_file, output_folder)


if __name__ == '__main__':
    test()
